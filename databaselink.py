from openpyxl import Workbook, load_workbook
from datetime import datetime


def append_to_excel(id, name, email, phno, dob, message,file_url):
    filename = "deeptumordata.xlsx"
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Email", "Phone Number", "DOB", "Timestamp", "Message", "Document Url"])

    row = [id, name, email, phno, dob, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), message, file_url]
    ws.append(row)
    wb.save(filename)
    print("appended")
